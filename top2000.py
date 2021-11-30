import requests
import openpyxl
from difflib import SequenceMatcher
import re
from pathlib import Path


class TOP2000:
    def __init__(self):
        self.full_list: list = []
        self.songs_radio2_api: list = []
        self.current_song_radio2_api: dict = {}
        self.current_song: dict = {}
        self.parse_top2000_list()

    def update_current_song(self):
        #Requests song currently playing
        response = requests.get("https://www.nporadio2.nl/api/tracks")
        
        data = response.json()
        
        self.songs_radio2_api = data['data']
        self.current_song_radio2_api = {'title': data['data'][0]['title'], 'artist': data['data'][0]['artist'], 'thumbnail': data['data'][0]['image_url_400x400']}

    def parse_top2000_list(self):
        xlsx_file = Path('lijsten/TOP-2000-2020.xlsx')
        wb_obj = openpyxl.load_workbook(xlsx_file)

        # Read the active sheet:
        sheet = wb_obj.active
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i != 0:
                self.full_list.append({'number': row[0], 'artist': row[1], 'title': row[2], 'year': row[3]})

    def get_current_song(self) -> dict:
        self.update_current_song()

        for song in self.full_list:
            acc_artist = float(SequenceMatcher(a=song['artist'], b=self.current_song_radio2_api['artist']).ratio())
            acc_title = float(SequenceMatcher(a=song['title'], b=self.current_song_radio2_api['title']).ratio())
            # if acc_artist > 0.8 and acc_title > 0.8:
            #     self.current_song = song
            #     self.current_song['image_url'] = self.current_song_radio2_api['image_url_400x400']
            #     return self.current_song
            # else:
            return self.current_song_radio2_api

    # Function search what songs are comming with the current song
    def get_future_song(self, amount=5):
        songs: list = []
        current_song = self.get_current_song()
        index: int = self.full_list.index(current_song) - 1
        for i in range(index, index - amount, -1):
            songs.append(self.full_list[i])
        return songs


# # Update the current song
# def update():
#     global data, titels, artiest, current_song_title, current_song_artist
#     response = requests.get("https://www.nporadio2.nl/api/tracks")

#     data = response.json()

#     current_song = data['data'][0]
#     current_song_title = current_song['title']
#     current_song_artist = current_song['artist']

#     xlsx_file = Path('TOP-2000-2020.xlsx')
#     wb_obj = openpyxl.load_workbook(xlsx_file)

#     # Read the active sheet:
#     sheet = wb_obj.active
#     data = {}

#     for i, row in enumerate(sheet.iter_rows(values_only=True)):
#         if i == 0:
#             data[row[0]] = []
#             data[row[1]] = []
#             data[row[2]] = []
#             data[row[3]] = []

#         else:
#             data['NR.'].append(row[0])
#             data['ARTIEST'].append(row[1])
#             data['TITEL'].append(row[2])
#             data['JAAR'].append(row[3])

#     titels = data['TITEL']
#     artiest = data['ARTIEST']

#     # Function to return the information of the current playing song according to radio 2 api
# def get_current_song():
#     update()
#     time.sleep(1)
#     indexes_title = get_index_positions(titels, current_song_title)
#     indexes_artist = get_index_positions(artiest, current_song_artist)
#     now_playing = {}

#     if len(indexes_title) == 1:
#         now_playing["Nummer"] = data['NR.'][indexes_title[0]]
#         now_playing["Artiest"] = data['ARTIEST'][indexes_title[0]]
#         now_playing["Titel"] = data['TITEL'][indexes_title[0]]
#         now_playing["Jaar"] = data['JAAR'][indexes_title[0]]
#     elif len(indexes_artist) == 1:
#         now_playing["Nummer"] = data['NR.'][indexes_artist[0]]
#         now_playing["Artiest"] = data['ARTIEST'][indexes_artist[0]]
#         now_playing["Titel"] = data['TITEL'][indexes_artist[0]]
#         now_playing["Jaar"] = data['JAAR'][indexes_artist[0]]
#     else:
#         for nummer in indexes_title:
#             if nummer in indexes_artist:
#                 now_playing["Nummer"] = data['NR.'][nummer]
#                 now_playing["Artiest"] = data['ARTIEST'][nummer]
#                 now_playing["Titel"] = data['TITEL'][nummer]
#                 now_playing["Jaar"] = data['JAAR'][nummer]

#     return now_playing

# # Get a list of indexes for posistions in list of the Top2000
# def get_index_positions(list_of_elems, element):
#     index_pos_list = []
#     index_pos = 0
#     # Parse items between brackets because radio 2 api is weird like that
#     element = re.split(r'\(.*?\)', element)
#     element = element[0]
#     while True:
#         try:
#             # Search for item in list from indexPos to the end of list
#             index_pos = list_of_elems.index(element, index_pos)
#             # Add the index position in list
#             index_pos_list.append(index_pos)
#             index_pos += 1
#         except ValueError as e:
#             break

#     # Check with a different method if first one doesn't work
#     if len(index_pos_list) == 0:
#         for elem in list_of_elems:
#             if element in elem or elem in element:
#                 pos = list_of_elems.index(elem)
#                 index_pos_list.append(pos)

#     return index_pos_list

# # Function to return the information of the current playing song according to radio 2 api
# def get_current_song():
#     update()
#     time.sleep(1)
#     indexes_title = get_index_positions(titels, current_song_title)
#     indexes_artist = get_index_positions(artiest, current_song_artist)
#     now_playing = {}

#     if len(indexes_title) == 1:
#         now_playing["Nummer"] = data['NR.'][indexes_title[0]]
#         now_playing["Artiest"] = data['ARTIEST'][indexes_title[0]]
#         now_playing["Titel"] = data['TITEL'][indexes_title[0]]
#         now_playing["Jaar"] = data['JAAR'][indexes_title[0]]
#     elif len(indexes_artist) == 1:
#         now_playing["Nummer"] = data['NR.'][indexes_artist[0]]
#         now_playing["Artiest"] = data['ARTIEST'][indexes_artist[0]]
#         now_playing["Titel"] = data['TITEL'][indexes_artist[0]]
#         now_playing["Jaar"] = data['JAAR'][indexes_artist[0]]
#     else:
#         for nummer in indexes_title:
#             if nummer in indexes_artist:
#                 now_playing["Nummer"] = data['NR.'][nummer]
#                 now_playing["Artiest"] = data['ARTIEST'][nummer]
#                 now_playing["Titel"] = data['TITEL'][nummer]
#                 now_playing["Jaar"] = data['JAAR'][nummer]

#     return now_playing

# # Function search what songs are comming with the current song
# def get_future_song(int=5):
#     songs = []
#     current_song = get_current_song()
#     nummer = current_song['Nummer'] - 2
#     for i in range(nummer, nummer - int, -1):
#         item = {}
#         item["Nummer"] = data['NR.'][i]
#         item["Artiest"] = data['ARTIEST'][i]
#         item["Titel"] = data['TITEL'][i]
#         item["Jaar"] = data['JAAR'][i]
#         songs.append(item)
#     return songs
